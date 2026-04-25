# -*- coding: utf-8 -*-
import argparse
import csv
import os
import re
from datetime import datetime


try:
    import openpyxl
except ImportError:  # pragma: no cover - optional dependency at runtime
    openpyxl = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency at runtime
    xlrd = None


ALIASES = {
    "default_code": [
        "default_code",
        "code",
        "编码",
        "物料编码",
        "料号",
        "sku",
        "产品编码",
        "型号编码",
    ],
    "name": ["name", "名称", "品名", "材料名称", "产品名称", "物料名称"],
    "spec": ["spec", "规格", "规格型号", "型号", "material_spec"],
    "category": ["category", "分类", "材料分类", "品类", "三级分类", "原材料分类"],
    "category_path": ["category_path", "分类路径", "分类全路径", "类目路径"],
    "brand": ["brand", "品牌", "品牌名称"],
    "color": ["color", "颜色", "色系"],
    "main_vendor": ["main_vendor", "供应商", "主要供应商", "厂商", "供应厂商", "vendor"],
    "manufacturer": ["manufacturer", "制造商", "生产厂家", "厂家"],
    "width": ["width", "宽", "宽度", "门幅"],
    "length": ["length", "长", "长度", "卷长"],
    "length_m": ["length_m", "长度(m)", "长度m", "长(m)", "卷长(m)"],
    "length_mm": ["length_mm", "长度(mm)", "长度mm", "长(mm)"],
    "thickness": ["thickness", "厚度", "总厚度", "厚度mm"],
    "rs_type": ["rs_type", "r/s", "卷片", "卷片类型", "形态", "卷料/片料"],
    "raw_material_price_m2": [
        "raw_material_price_m2",
        "price_per_m2",
        "单价/m²",
        "单价/m2",
        "平米价",
        "每平米价格",
        "rmb/m²",
        "rmb/m2",
        "价格/m²",
        "价格/m2",
    ],
    "weight_gram": ["weight_gram", "克重", "重量", "克重(g)", "g数"],
    "density": ["density", "密度", "密度(g/cm3)", "密度(g/cm³)"],
    "material_type": ["material_type", "材质", "材质/牌号", "材质牌号"],
    "origin": ["origin", "产地", "原产地"],
    "track_batch": ["track_batch", "批次管理", "批次追踪", "lot_tracking"],
    "min_order_qty": ["min_order_qty", "moq", "起订量", "最小起订量", "最小采购量"],
    "lead_time": ["lead_time", "交期", "采购周期", "交货期", "leadtime"],
    "purchase_uom": ["purchase_uom", "采购单位", "单位"],
    "price_unit": ["price_unit", "价格单位", "报价单位"],
    "contact_info": ["contact_info", "联系方式", "联系人电话"],
    "incoterms": ["incoterms", "贸易条款", "贸易方式"],
    "quote_date": ["quote_date", "报价日期", "报价时间"],
    "application": ["application", "应用", "应用场景", "用途"],
    "process_note": ["process_note", "工艺说明", "加工说明", "加工工艺说明"],
    "caution": ["caution", "注意事项", "备注", "说明"],
}

REQUIRED_HEADER_CANDIDATES = {"name", "default_code", "spec", "main_vendor", "category"}
TEXT_NULLS = {"", "none", "null", "nan", "n/a", "na", "-"}
TRUE_VALUES = {"1", "true", "yes", "y", "是", "有", "需", "开"}
FALSE_VALUES = {"0", "false", "no", "n", "否", "无", "不", "关"}
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m", "%Y/%m", "%Y.%m")

LEGACY_SKIPPED_SHEETS = {"油墨", "参考价格"}
LEGACY_EXCHANGE_RATES = {
    "RMB": 1.0,
    "HKD": 6.4 / 7.75,
    "USD": 6.4,
    "JPY": 6.4 / 99.0,
}
LEGACY_CATEGORY_XMLIDS = {
    "PET": "diecut.category_other_auxiliary",
    "PP": "diecut.category_other_auxiliary",
    "PC": "diecut.category_other_auxiliary",
    "PMMA": "diecut.category_other_auxiliary",
    "PORON": "diecut.category_foam",
    "スポンジ": "diecut.category_foam",
    "海绵": "diecut.category_foam",
    "海綿": "diecut.category_foam",
    "ゴム": "diecut.category_foam",
    "不织布": "diecut.category_foam",
    "両面テープ": "diecut.category_tape",
    "片面テープ": "diecut.category_tape",
    "単面テープ": "diecut.category_tape",
    "テープ": "diecut.category_tape",
    "アルミテープ": "diecut.category_tape",
    "デープ": "diecut.category_tape",
    "离型PET": "diecut.category_release_film",
    "离形PET": "diecut.category_release_film",
    "セパ": "diecut.category_release_film",
    "保護ファイム": "diecut.category_protection_film",
    "保護フィルム": "diecut.category_protection_film",
    "保護フィムル": "diecut.category_protection_film",
    "保护膜": "diecut.category_protection_film",
    "微粘フィムル": "diecut.category_protection_film",
    "保護膜": "diecut.category_protection_film",
    "防曝膜": "diecut.category_protection_film",
    "放熱シート": "diecut.category_graphite",
    "绝缘纸": "diecut.category_paper",
    "アルミ": "diecut.category_metal_foil",
    "銅＋PET": "diecut.category_metal_foil",
    "电磁波吸收": "diecut.category_shielding",
    "MYLAR MO": "diecut.category_other_auxiliary",
    "UPE": "diecut.category_other_auxiliary",
    "ABS": "diecut.category_other_auxiliary",
    "Glass": "diecut.category_other_auxiliary",
    "シリコン": "diecut.category_other_auxiliary",
    "PVC": "diecut.category_other_auxiliary",
    "耐高温PET": "diecut.category_protection_film",
    "外注品": "diecut.category_other_auxiliary",
    "膠袋": "diecut.category_other_auxiliary",
    "外箱": "diecut.category_other_auxiliary",
    "消耗品": "diecut.category_other_auxiliary",
    "平卡": "diecut.category_other_auxiliary",
    "隔板": "diecut.category_other_auxiliary",
    "4刀卡": "diecut.category_other_auxiliary",
    "魔术贴": "diecut.category_other_auxiliary",
    "材料": "diecut.category_other_auxiliary",
}


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().replace("\r", "").replace("\u3000", " ")
    return str(value).strip()


def normalize_header(value):
    text = clean_text(value)
    if not text:
        return ""
    compact = text.lower()
    compact = compact.replace("（", "(").replace("）", ")")
    compact = compact.replace("㎡", "m2").replace("m²", "m2")
    compact = compact.replace("/", "").replace("_", "").replace("-", "")
    compact = compact.replace(" ", "")
    compact = compact.replace("(", "").replace(")", "")
    return compact


ALIAS_LOOKUP = {
    normalize_header(alias): field_name
    for field_name, aliases in ALIASES.items()
    for alias in aliases
}


def parse_bool(value):
    text = clean_text(value).lower()
    if not text:
        return None
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def parse_int(value):
    number = parse_number(value)
    return int(number) if number is not None else None


def parse_date(value):
    text = clean_text(value)
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def detect_header_row(rows):
    best_index = None
    best_score = -1
    for index, row in enumerate(rows[:10]):
        score = 0
        for cell in row:
            if normalize_header(cell) in ALIAS_LOOKUP:
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    if best_index is None or best_score < 2:
        raise ValueError("无法识别表头，请确认文件包含名称、分类、供应商等列。")
    return best_index


def build_header_map(header_row):
    header_map = {}
    seen = set()
    for idx, cell in enumerate(header_row):
        canonical = ALIAS_LOOKUP.get(normalize_header(cell))
        if canonical and canonical not in seen:
            header_map[canonical] = idx
            seen.add(canonical)
    if not (seen & REQUIRED_HEADER_CANDIDATES):
        raise ValueError("表头未识别到关键列，请至少提供名称、编码、规格、分类、供应商中的部分列。")
    return header_map


def iter_csv_rows(path):
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"无法解析 CSV 编码: {path}") from last_error


def iter_xlsx_rows(path, sheet_name=None):
    if openpyxl is None:
        raise ValueError("当前环境未安装 openpyxl，无法读取 xlsx 文件。请改用 csv 或安装 openpyxl。")
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"未找到工作表: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = next((ws for ws in workbook.worksheets if ws.max_row and ws.max_column), None)
            if sheet is None:
                raise ValueError("工作簿没有可用数据表。")
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def iter_xls_rows(path, sheet_name=None):
    if xlrd is None:
        raise ValueError("当前环境未安装 xlrd，无法读取 xls 文件。")
    workbook = xlrd.open_workbook(path)
    if sheet_name:
        if sheet_name not in workbook.sheet_names():
            raise ValueError(f"未找到工作表: {sheet_name}")
        sheet = workbook.sheet_by_name(sheet_name)
    else:
        sheet = next((ws for ws in workbook.sheets() if ws.nrows and ws.ncols), None)
        if sheet is None:
            raise ValueError("工作簿没有可用数据表。")
    return [sheet.row_values(index) for index in range(sheet.nrows)]


def infer_rs_type_and_length(row):
    explicit_rs_type = clean_text(row.get("rs_type")).upper()
    if explicit_rs_type in {"R", "S"}:
        rs_type = explicit_rs_type
    else:
        value_text = clean_text(row.get("rs_type")) or clean_text(row.get("length")) or clean_text(row.get("name"))
        rs_type = "S" if "片" in value_text else "R"

    if row.get("length_m") is not None:
        return rs_type, parse_number(row.get("length_m"))
    if row.get("length_mm") is not None:
        length_mm = parse_number(row.get("length_mm"))
        return "S", (length_mm / 1000.0 if length_mm is not None else None)

    raw_length = row.get("length")
    text = clean_text(raw_length).lower()
    number = parse_number(raw_length)
    if number is None:
        return rs_type, None
    if "mm" in text:
        return "S", number / 1000.0
    if "cm" in text:
        return rs_type, number / 100.0
    if "米" in text or text.endswith("m") or "m" in text:
        return rs_type, number
    if rs_type == "S" and number > 10:
        return rs_type, number / 1000.0
    return rs_type, number


def normalize_record(record):
    name = clean_text(record.get("name"))
    default_code = clean_text(record.get("default_code"))
    spec = clean_text(record.get("spec"))
    category = clean_text(record.get("category"))
    category_path = clean_text(record.get("category_path"))
    if not any([name, default_code, spec]):
        return None

    rs_type, length_m = infer_rs_type_and_length(record)
    return {
        "_source_row": record["_source_row"],
        "default_code": default_code or False,
        "name": name or spec or default_code,
        "spec": spec or False,
        "category": category or False,
        "category_path": category_path or False,
        "brand": clean_text(record.get("brand")) or False,
        "color": clean_text(record.get("color")) or False,
        "main_vendor": clean_text(record.get("main_vendor")) or False,
        "manufacturer": clean_text(record.get("manufacturer")) or False,
        "width": parse_number(record.get("width")),
        "length": length_m,
        "thickness": parse_number(record.get("thickness")),
        "rs_type": rs_type,
        "raw_material_price_m2": parse_number(record.get("raw_material_price_m2")),
        "weight_gram": parse_number(record.get("weight_gram")),
        "density": parse_number(record.get("density")),
        "material_type": clean_text(record.get("material_type")) or False,
        "origin": clean_text(record.get("origin")) or False,
        "track_batch": parse_bool(record.get("track_batch")),
        "min_order_qty": parse_number(record.get("min_order_qty")),
        "lead_time": parse_int(record.get("lead_time")),
        "purchase_uom": clean_text(record.get("purchase_uom")) or False,
        "price_unit": clean_text(record.get("price_unit")) or False,
        "contact_info": clean_text(record.get("contact_info")) or False,
        "incoterms": clean_text(record.get("incoterms")) or False,
        "quote_date": parse_date(record.get("quote_date")),
        "application": clean_text(record.get("application")) or False,
        "process_note": clean_text(record.get("process_note")) or False,
        "caution": clean_text(record.get("caution")) or False,
    }


def load_generic_rows(path, sheet_name=None):
    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        rows = iter_csv_rows(path)
    elif extension == ".xlsx":
        rows = iter_xlsx_rows(path, sheet_name=sheet_name)
    elif extension == ".xls":
        rows = iter_xls_rows(path, sheet_name=sheet_name)
    else:
        raise ValueError("仅支持 .xls、.xlsx 或 .csv 文件。")

    if not rows:
        return []

    header_index = detect_header_row(rows)
    header_map = build_header_map(rows[header_index])
    records = []
    for source_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(clean_text(cell) for cell in row):
            continue
        record = {"_source_row": source_index}
        for field_name, col_index in header_map.items():
            record[field_name] = row[col_index] if col_index < len(row) else None
        normalized = normalize_record(record)
        if normalized:
            records.append(normalized)
    return records


def is_legacy_material_workbook(path):
    if xlrd is None or os.path.splitext(path)[1].lower() != ".xls":
        return False
    workbook = xlrd.open_workbook(path)
    for sheet in workbook.sheets():
        for row_index in range(min(sheet.nrows, 4)):
            row_text = " ".join(clean_text(value) for value in sheet.row_values(row_index))
            if "材料品番" in row_text and "購入先" in row_text:
                return True
    return False


def excel_date_to_date(value, datemode):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and value > 1000:
        return xlrd.xldate_as_datetime(value, datemode).date()
    return parse_date(value)


def pick_legacy_purchase_price(row):
    price_columns = [
        ("RMB", 16),
        ("HKD", 15),
        ("USD", 17),
        ("JPY", 18),
    ]
    for currency, index in price_columns:
        if index >= len(row):
            continue
        amount = parse_number(row[index])
        if amount is not None and amount > 0:
            return amount * LEGACY_EXCHANGE_RATES[currency]
    return None


def compute_legacy_area_m2(width_mm, length_value, length_unit, rs_type):
    if width_mm is None or length_value is None:
        return None
    width_m = width_mm / 1000.0
    normalized_unit = clean_text(length_unit).lower()
    normalized_rs_type = clean_text(rs_type).upper()
    if normalized_unit == "mm" or normalized_rs_type == "S":
        length_m = length_value / 1000.0
    else:
        length_m = length_value
    area = width_m * length_m
    return area if area > 0 else None


def normalize_legacy_category_label(value):
    return clean_text(value).replace("　", " ").strip()


def build_legacy_record(workbook, sheet, row_index):
    row = sheet.row_values(row_index)
    default_code = clean_text(row[4] if len(row) > 4 else "")
    if not default_code:
        return None

    material_type = normalize_legacy_category_label(row[5] if len(row) > 5 else "")
    if not material_type:
        raise ValueError("材料種類为空，无法分类")

    price_total_rmb = pick_legacy_purchase_price(row)
    if price_total_rmb is None:
        raise ValueError("未找到可用采购价格")

    width_mm = parse_number(row[8] if len(row) > 8 else None)
    length_value = parse_number(row[9] if len(row) > 9 else None)
    length_unit = clean_text(row[10] if len(row) > 10 else "")
    rs_type = clean_text(row[11] if len(row) > 11 else "")
    area_m2 = compute_legacy_area_m2(width_mm, length_value, length_unit, rs_type)
    if area_m2 is None:
        raise ValueError("缺少可折算面积的尺寸信息")

    raw_material_price_m2 = price_total_rmb / area_m2
    thickness_value = parse_number(row[7] if len(row) > 7 else None)
    quote_date = excel_date_to_date(row[0] if len(row) > 0 else None, workbook.datemode)
    lead_time = clean_text(row[20] if len(row) > 20 else "")
    name = f"{default_code} {material_type}".strip()

    return {
        "_source_row": f"{sheet.name}:{row_index + 1}",
        "default_code": default_code,
        "name": name,
        "spec": default_code,
        "category": material_type,
        "category_path": False,
        "brand": False,
        "color": clean_text(row[6] if len(row) > 6 else "") or False,
        "main_vendor": clean_text(row[1] if len(row) > 1 else "") or False,
        "manufacturer": False,
        "width": width_mm,
        "length": area_m2 / (width_mm / 1000.0) if width_mm else None,
        "thickness": thickness_value,
        "rs_type": clean_text(rs_type).upper() if clean_text(rs_type).upper() in {"R", "S"} else ("S" if clean_text(length_unit).lower() == "mm" else "R"),
        "raw_material_price_m2": raw_material_price_m2,
        "weight_gram": None,
        "density": None,
        "material_type": material_type,
        "origin": False,
        "track_batch": None,
        "min_order_qty": parse_number(row[19] if len(row) > 19 else None),
        "lead_time": parse_int(lead_time),
        "purchase_uom": clean_text(row[13] if len(row) > 13 else "") or False,
        "price_unit": clean_text(row[13] if len(row) > 13 else "") or False,
        "contact_info": False,
        "incoterms": clean_text(row[2] if len(row) > 2 else "") or False,
        "quote_date": quote_date,
        "application": False,
        "process_note": False,
        "caution": f"导入自工作表 {sheet.name}",
        "_legacy_sheet_name": sheet.name,
        "_legacy_material_type": material_type,
    }


def load_legacy_material_rows(path, sheet_name=None):
    if xlrd is None:
        raise ValueError("当前环境未安装 xlrd，无法读取 xls 文件。")

    workbook = xlrd.open_workbook(path)
    if sheet_name:
        selected_sheets = [workbook.sheet_by_name(sheet_name)]
    else:
        selected_sheets = [sheet for sheet in workbook.sheets() if sheet.name not in LEGACY_SKIPPED_SHEETS]

    records = []
    for sheet in selected_sheets:
        header_row_index = None
        for idx in range(min(sheet.nrows, 5)):
            row_text = " ".join(clean_text(value) for value in sheet.row_values(idx))
            if "材料品番" in row_text and "購入先" in row_text:
                header_row_index = idx
                break
        if header_row_index is None:
            continue

        data_start = header_row_index + 2
        for row_index in range(data_start, sheet.nrows):
            if not any(clean_text(cell) for cell in sheet.row_values(row_index)):
                continue
            default_code = clean_text(sheet.cell_value(row_index, 4)) if sheet.ncols > 4 else ""
            try:
                record = build_legacy_record(workbook, sheet, row_index)
            except Exception as exc:
                records.append(
                    {
                        "_source_row": f"{sheet.name}:{row_index + 1}",
                        "default_code": default_code or False,
                        "name": default_code or False,
                        "_load_error": str(exc),
                    }
                )
                continue
            if record:
                records.append(record)
    return records


def load_rows(path, sheet_name=None):
    if is_legacy_material_workbook(path):
        return load_legacy_material_rows(path, sheet_name=sheet_name)
    return load_generic_rows(path, sheet_name=sheet_name)


class RawMaterialImporter:
    def __init__(self, env):
        self.env = env
        self.brand_model = env["diecut.brand"].sudo()
        self.color_model = env["diecut.color"].sudo()
        self.partner_model = env["res.partner"].sudo()
        self.category_model = env["product.category"].sudo()
        self.product_model = env["product.template"].sudo()

    def import_file(self, file_path, sheet_name=None, dry_run=False, fail_output_path=None):
        records = load_rows(file_path, sheet_name=sheet_name)
        summary = {
            "total": len(records),
            "success": 0,
            "created": 0,
            "updated": 0,
            "failed": 0,
            "failures": [],
        }
        for record in records:
            try:
                result = self.import_record(record, dry_run=dry_run)
                summary["success"] += 1
                if result == "created":
                    summary["created"] += 1
                else:
                    summary["updated"] += 1
            except Exception as exc:  # pragma: no cover - exercised via Odoo runtime
                summary["failed"] += 1
                summary["failures"].append(
                    {
                        "row_number": record["_source_row"],
                        "material_label": record.get("default_code") or record.get("name") or "",
                        "reason": str(exc),
                    }
                )
        if fail_output_path:
            self.write_failures(summary["failures"], fail_output_path)
        return summary

    def import_record(self, record, dry_run=False):
        if record.get("_load_error"):
            raise ValueError(record["_load_error"])
        values = self.build_product_values(record)
        product = self.find_existing_product(record, values.get("main_vendor_id"))
        action = "updated" if product else "created"
        if dry_run:
            return action
        if product:
            product.write(values)
        else:
            self.product_model.create(values)
        return action

    def build_product_values(self, record):
        values = {
            "is_raw_material": True,
            "purchase_ok": True,
            "sale_ok": False,
        }
        field_mapping = [
            "default_code",
            "name",
            "spec",
            "width",
            "length",
            "thickness",
            "rs_type",
            "raw_material_price_m2",
            "weight_gram",
            "density",
            "material_type",
            "origin",
            "min_order_qty",
            "lead_time",
            "purchase_uom",
            "price_unit",
            "contact_info",
            "incoterms",
            "quote_date",
            "application",
            "process_note",
            "caution",
        ]
        for field_name in field_mapping:
            value = record.get(field_name)
            if value not in (None, False, ""):
                values[field_name] = value
        if record.get("track_batch") is not None:
            values["track_batch"] = record["track_batch"]

        category = self.resolve_category(record)
        values["categ_id"] = category.id
        if record.get("brand"):
            values["brand_id"] = self.resolve_brand(record["brand"]).id
        if record.get("color"):
            values["color_id"] = self.resolve_color(record["color"]).id
        if record.get("main_vendor"):
            values["main_vendor_id"] = self.resolve_partner(record["main_vendor"], supplier=True).id
        if record.get("manufacturer"):
            values["manufacturer_id"] = self.resolve_partner(record["manufacturer"], supplier=False).id
        return values

    def find_existing_product(self, record, main_vendor_id):
        if record.get("default_code"):
            product = self.product_model.search(
                [("is_raw_material", "=", True), ("default_code", "=", record["default_code"])],
                limit=1,
            )
            if product:
                return product
        domain = [
            ("is_raw_material", "=", True),
            ("name", "=", record.get("name") or False),
            ("spec", "=", record.get("spec") or False),
        ]
        if main_vendor_id:
            domain.append(("main_vendor_id", "=", main_vendor_id))
        else:
            domain.append(("main_vendor_id", "=", False))
        return self.product_model.search(domain, limit=1)

    def resolve_brand(self, name):
        brand = self.brand_model.search([("name", "=", name)], limit=1)
        return brand or self.brand_model.create({"name": name})

    def resolve_color(self, name):
        color = self.color_model.search([("name", "=", name)], limit=1)
        return color or self.color_model.create({"name": name})

    def resolve_partner(self, name, supplier=False):
        partner = self.partner_model.search(
            ["|", ("name", "=", name), ("short_name", "=", name)],
            limit=1,
        )
        if partner:
            if supplier and partner.supplier_rank <= 0:
                partner.supplier_rank = 1
            return partner
        values = {"name": name, "is_company": True}
        if supplier:
            values["supplier_rank"] = 1
        return self.partner_model.create(values)

    def resolve_category(self, record):
        legacy_xmlid = LEGACY_CATEGORY_XMLIDS.get(clean_text(record.get("_legacy_material_type")))
        if legacy_xmlid:
            return self.env.ref(legacy_xmlid)

        candidates = []
        if record.get("category_path"):
            candidates.extend(self.build_category_path_candidates(record["category_path"]))
        if record.get("category"):
            candidates.append(record["category"])

        for candidate in candidates:
            matched = self.match_category(candidate)
            if matched:
                return matched

        label = record.get("category_path") or record.get("category") or ""
        raise ValueError(f"未识别分类：{label}")

    def build_category_path_candidates(self, category_path):
        text = clean_text(category_path).replace("\\", "/").replace(">", "/").replace("／", "/")
        parts = [segment.strip() for segment in text.split("/") if segment.strip()]
        candidates = []
        if parts:
            candidates.append(" / ".join(parts))
            candidates.append("/".join(parts))
            candidates.extend(reversed(parts))
        return candidates

    def match_category(self, label):
        if not label:
            return False
        normalized_target = self.normalize_category_label(label)
        raw_categories = self.category_model.search([("category_type", "=", "raw")])
        exact_matches = raw_categories.filtered(
            lambda categ: self.normalize_category_label(categ.complete_name) == normalized_target
            or self.normalize_category_label(categ.name) == normalized_target
        )
        if len(exact_matches) == 1:
            return exact_matches[0]
        leaf_matches = raw_categories.filtered(lambda categ: self.normalize_category_label(categ.name) == normalized_target)
        return leaf_matches[0] if len(leaf_matches) == 1 else False

    @staticmethod
    def normalize_category_label(value):
        return clean_text(value).replace(">", "/").replace("\\", "/").replace("／", "/").replace(" / ", "/")

    @staticmethod
    def write_failures(failures, output_path):
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["row_number", "material_label", "reason"])
            writer.writeheader()
            writer.writerows(failures)


def format_summary(summary):
    return (
        f"总行数: {summary['total']}\n"
        f"成功: {summary['success']}\n"
        f"新增: {summary['created']}\n"
        f"更新: {summary['updated']}\n"
        f"失败: {summary['failed']}"
    )


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="导入原材料 xls/xlsx/csv 到 Odoo 原材料库。")
    parser.add_argument("--file", required=True, help="xls、xlsx 或 csv 文件路径")
    parser.add_argument("--sheet", help="仅导入指定工作表")
    parser.add_argument("--dry-run", action="store_true", help="仅解析和匹配，不写入数据库")
    parser.add_argument("--fail-output", help="失败明细 CSV 输出路径")
    parser.add_argument("--config", help="Odoo 配置文件路径")
    parser.add_argument("--database", help="数据库名")
    parser.add_argument("--db-host", help="数据库主机")
    parser.add_argument("--db-user", help="数据库用户")
    parser.add_argument("--db-password", help="数据库密码")
    return parser.parse_args(argv)


def build_odoo_env(config_path, database, db_host=None, db_user=None, db_password=None):
    import odoo
    from odoo import SUPERUSER_ID, api
    from odoo.modules.registry import Registry
    from odoo.service.server import load_server_wide_modules

    if not config_path or not database:
        raise ValueError("非测试环境运行时必须提供 --config 和 --database。")

    odoo.tools.config.parse_config(["-c", config_path, "-d", database])
    if db_host:
        odoo.tools.config["db_host"] = db_host
    if db_user:
        odoo.tools.config["db_user"] = db_user
    if db_password:
        odoo.tools.config["db_password"] = db_password
    load_server_wide_modules()
    registry = Registry(database)
    cursor = registry.cursor()
    env = api.Environment(cursor, SUPERUSER_ID, {})
    return env, cursor


def run_cli(argv=None):
    args = parse_args(argv)
    env = globals().get("env")
    cursor = None
    if env is None:
        env, cursor = build_odoo_env(
            args.config,
            args.database,
            db_host=args.db_host,
            db_user=args.db_user,
            db_password=args.db_password,
        )

    try:
        importer = RawMaterialImporter(env)
        summary = importer.import_file(
            file_path=args.file,
            sheet_name=args.sheet,
            dry_run=args.dry_run,
            fail_output_path=args.fail_output,
        )
        if cursor:
            if args.dry_run:
                cursor.rollback()
            else:
                cursor.commit()
        print(format_summary(summary))
        if args.fail_output and summary["failed"]:
            print(f"失败明细: {args.fail_output}")
        return summary
    finally:
        if cursor:
            cursor.close()


if __name__ == "__main__":
    run_cli()
